import os, re
from zipfile import ZipFile
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from bs4 import BeautifulSoup

# ====== Các biểu thức chính quy tìm thông tin nhạy cảm ======
email_regex = r"[a-z0-9]+[\._]?[a-z0-9]+@[\w]+\.[\w]{2,3}"                  # Email
phone_regex = r"(0\d{9})|(\(\+84\)\d{2} \d{3} \d{4})"                       # SĐT Việt Nam
ssn_regex = r"\b\d{9}\b|\b\d{12}\b"                                        # CCCD 9 hoặc 12 số
regexes = {
    "Email": email_regex,
    "Phone": phone_regex,
    "CCCD": ssn_regex
}

# ====== Tìm kiếm thông tin nhạy cảm trong văn bản ======
def find_pii(text):
    matches = []
    for label, pattern in regexes.items():
        for match in re.findall(pattern, text):
            if isinstance(match, tuple):  # Với regex nhóm, lấy phần không rỗng
                match = [m for m in match if m][0]
            matches.append((label, match))
    return matches

# ====== In kết quả tìm được nếu có ======
def print_matches(filepath, matches):
    if matches:
        print(f"\nFile: {filepath}")
        for label, match in matches:
            print(f"  {label}: {match}")

# ====== Xử lý file TXT, CSV, PY ======
def parse_text_file(filepath):
    with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
        matches = find_pii(f.read())
    print_matches(filepath, matches)

# ====== Xử lý file DOCX (file Word) ======
def parse_docx(filepath):
    with ZipFile(filepath, "r") as z:
        data = z.read("word/document.xml").decode("utf-8")
        matches = find_pii(data)
    print_matches(filepath, matches)

# ====== Xử lý file PDF ======
def parse_pdf(filepath):
    text = ""
    with open(filepath, "rb") as f:
        pdf = PdfReader(f)
        for page in pdf.pages:
            text += page.extract_text() or ""
    print_matches(filepath, find_pii(text))

# ====== Xử lý file Excel (.xls, .xlsx) ======
def parse_excel(filepath):
    text = ""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        for row in wb[sheet].iter_rows(values_only=True):
            text += " ".join([str(cell) for cell in row if cell]) + "\n"
    print_matches(filepath, find_pii(text))

# ====== Xử lý file HTML ======
def parse_html(filepath):
    with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
        soup = BeautifulSoup(f, "html.parser")
        text = soup.get_text()
    print_matches(filepath, find_pii(text))

# ====== Duyệt tất cả file trong thư mục chỉ định và gọi hàm xử lý phù hợp ======
def scan_directory(directory):
    print(f"Scanning folder: {directory}")
    for root, _, files in os.walk(directory):
        for file in files:
            filepath = os.path.join(root, file)
            try:
                if file.endswith((".txt", ".csv", ".py")):
                    parse_text_file(filepath)
                elif file.endswith(".docx"):
                    parse_docx(filepath)
                elif file.endswith(".pdf"):
                    parse_pdf(filepath)
                elif file.endswith((".xls", ".xlsx")):
                    parse_excel(filepath)
                elif file.endswith((".html", ".htm")):
                    parse_html(filepath)
            except Exception as e:
                print(f" Error reading {file}: {e}")

# ====== CHỈNH ĐƯỜNG DẪN THƯ MỤC TẠI ĐÂY ======
if __name__ == "__main__":
    target_dir = "C:\\Users\\Phi Hung\\Desktop\\LTATTT\\Performing Discovery"
    scan_directory(target_dir)
